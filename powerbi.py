from SOMcreator import Project
import json
import openpyxl

from constants import  FEHLER_PATH

def write_bauteilklass(proj,ws):
    ws.title = "bauteilKlassifikation"
    ws.cell(row=1,column=1).value = "bauteilKlassifikation"
    ws.cell(row=1,column=2).value = "Name"
    ws.cell(row=1,column=3).value = "AttributAnzahl"
    ident_dict = {obj.ident_value:obj.name for obj in proj.objects}
    ident_dict["aus"] = "Ausbau"
    ident_dict["lst"] = "Leit und Sicherungstechnik"
    ident_dict["lst.geo"] = "Leit und Sicherungstechnik GEO"

    ma= max(len(obj.ident_value.split(".")) for obj in proj.objects if obj.ident_value is not None and obj.ident_value)

    for i in range(ma):
        ws.cell(row=1,column=i+4).value = f"Level {i+1}"
    for i,obj in enumerate(proj.objects,start=2):
        ident = obj.ident_value
        #print(f"ident: {ident}")
        if ident is None:
            continue
        ws.cell(row=i,column=1).value = ident
        ws.cell(row=i,column=2).value = obj.name
        attrib_count =sum(len(pset.attributes) for pset in [property_set for property_set in obj.property_sets])
        ws.cell(row=i,column=3).value = attrib_count

        levels = ident.split(".")
        level = 0
        for level in range(1,ma):
            t = obj.name
            if level <len(levels):
                t = ident_dict[".".join(levels[:level])]
            ws.cell(row=i,column=level+3).value = t
        ws.cell(row=i,column=level+4).value = obj.name

def write_fehler_art(ws):
    ws.cell(1,1).value = "Status"
    ws.cell(1,2).value = "Text"
    with open(FEHLER_PATH,"r") as file:
        d = json.load(file)

    for row,(status,text) in enumerate(d.items(),start = 2):
        ws.cell(row,1).value = status
        ws.cell(row,2).value = text

def write_property_sets(proj:Project,ws):
    props = { prop.name:len(prop.attributes) for obj in proj.objects for prop in obj.property_sets }
    ws.cell(1,1).value = "PropertySet"
    ws.cell(1,2).value = "AttributAnzahl"

    for row,(name,count) in enumerate(props.items(),start = 2):
        ws.cell(row,1).value = name
        ws.cell(row,2).value = count

def handle_definitions(proj:Project,workbook_path:str):
    wb = openpyxl.Workbook()
    ws_bk = wb.active
    write_bauteilklass(proj,ws_bk)
    ws_fehler_art = wb.create_sheet("fehlerArt")
    write_fehler_art(ws_fehler_art)
    ws_property_sets = wb.create_sheet("propertySets")
    write_property_sets(proj,ws_property_sets)
    wb.save(workbook_path)
