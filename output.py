import sql
import sqlite3
import os
import openpyxl

HEADER = ["GUID","Beschreibung","Typ","PropertySet","Attribut","Datei","Bauteilklassifikation"]

def save_workbook(workbook,path):
    try:
        workbook.save(path)
    except PermissionError:
        print("-"*60)
        print(f"folgende Datei ist noch geöffnet: {path} \n Datei schließen und beliebige Taste Drücken")
        input("Achtung! Datei wird danach überschrieben!")
        save_workbook(workbook,path)

def create_issues(db_name,path):
    conn = sqlite3.connect(db_name)
    cursor = conn.cursor()
    dir = os.path.dirname(path)
    if not os.path.exists(dir):
        os.mkdir(dir)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    issues = sql.query_issues(cursor)
    for col_index,value in enumerate(HEADER,start=1):
        worksheet.cell(1,col_index,value)

    for row_index,column in enumerate(issues,start =2):
        for column_index,value in enumerate(column,start =1):
            if column_index == 1:
                value = value.encode('ascii', 'ignore').decode("utf-8")
            worksheet.cell(row_index,column_index,value)  #remove Whitespace

    save_workbook(workbook,path)


